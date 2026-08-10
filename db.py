# File: C:\code\SYT\db.py
"""烟酒店收银系统 — 数据层 (SQLite)。"""

import sqlite3
import os
import shutil
import sys
import datetime
import csv
import math
import logging
import json
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger("cashier.db")


class DatabaseError(Exception):
    """数据库业务异常。"""


# 示例商品（首次运行时空库填充）
SAMPLE_PRODUCTS = []


def _now() -> str:
    """返回当前 ISO-8601 时间戳（本地时间）。"""
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ------------------------------------------------------------------ #
class Database:
    """SQLite 数据库操作（上下文管理器连接复用）。"""

    def __init__(self, db_path: str = "", backup_dir: str = ""):
        self.db_path = db_path or os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "data", "cashier.db"
        )
        self.backup_dir = backup_dir or os.path.join(
            os.path.dirname(self.db_path), "..", "备份"
        )
        self.init_schema()

    # ------------------------------------------------------------------ #
    # 初始化 & 迁移
    # ------------------------------------------------------------------ #
    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA foreign_keys=ON")
        return conn

    def init_schema(self) -> None:
        os.makedirs(os.path.dirname(self.db_path), exist_ok=True)
        with self._connect() as conn:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS products (
                    id           INTEGER PRIMARY KEY AUTOINCREMENT,
                    barcode      TEXT,
                    name         TEXT NOT NULL,
                    category     TEXT DEFAULT '其他',
                    cost_price   REAL NOT NULL DEFAULT 0,
                    sell_price   REAL NOT NULL DEFAULT 0,
                    member_price REAL,
                    stock        INTEGER NOT NULL DEFAULT 0,
                    low_stock    INTEGER NOT NULL DEFAULT 10,
                    created_at   TEXT NOT NULL
                );
                """
            )
            # 迁移：去掉 barcode UNIQUE 约束，允许同名条码存不同商品
            try:
                ddl = conn.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='products'").fetchone()
                need_migrate = ddl and 'barcode TEXT UNIQUE' in (ddl[0] or '')
            except Exception:
                need_migrate = False
            if need_migrate:
                logger.info("迁移: 移除 products.barcode UNIQUE 约束")
                conn.executescript(
                    """
                    ALTER TABLE products RENAME TO products_old;
                    CREATE TABLE products (
                        id           INTEGER PRIMARY KEY AUTOINCREMENT,
                        barcode      TEXT,
                        name         TEXT NOT NULL,
                        category     TEXT DEFAULT '其他',
                        cost_price   REAL NOT NULL DEFAULT 0,
                        sell_price   REAL NOT NULL DEFAULT 0,
                        member_price REAL,
                        stock        INTEGER NOT NULL DEFAULT 0,
                        low_stock    INTEGER NOT NULL DEFAULT 10,
                        created_at   TEXT NOT NULL
                    );
                    INSERT INTO products SELECT * FROM products_old;
                    DROP TABLE products_old;
                    """
                )
            # 迁移：stock_logs 添加 cost_price 列
            try:
                conn.execute("ALTER TABLE stock_logs ADD COLUMN cost_price REAL")
            except sqlite3.OperationalError:
                pass  # 列已存在
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS sales (
                    id            INTEGER PRIMARY KEY AUTOINCREMENT,
                    total_amount  REAL NOT NULL DEFAULT 0,
                    discount      REAL NOT NULL DEFAULT 0,
                    paid_amount   REAL NOT NULL DEFAULT 0,
                    change_amount REAL NOT NULL DEFAULT 0,
                    member_id     INTEGER,
                    created_at    TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS sale_items (
                    id           INTEGER PRIMARY KEY AUTOINCREMENT,
                    sale_id      INTEGER NOT NULL,
                    product_id   INTEGER NOT NULL,
                    product_name TEXT NOT NULL,
                    price        REAL NOT NULL,
                    quantity     INTEGER NOT NULL,
                    subtotal     REAL NOT NULL,
                    FOREIGN KEY (sale_id) REFERENCES sales(id),
                    FOREIGN KEY (product_id) REFERENCES products(id)
                );
                CREATE TABLE IF NOT EXISTS stock_logs (
                    id           INTEGER PRIMARY KEY AUTOINCREMENT,
                    product_id   INTEGER NOT NULL,
                    product_name TEXT NOT NULL,
                    change_type  TEXT NOT NULL,
                    quantity     INTEGER NOT NULL,
                    stock_after  INTEGER NOT NULL,
                    cost_price   REAL,
                    note         TEXT DEFAULT '',
                    created_at   TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS members (
                    id         INTEGER PRIMARY KEY AUTOINCREMENT,
                    phone      TEXT UNIQUE,
                    name       TEXT NOT NULL DEFAULT '',
                    gender     TEXT NOT NULL DEFAULT '',
                    created_at TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS product_operation_logs (
                    id             INTEGER PRIMARY KEY AUTOINCREMENT,
                    operation_type TEXT NOT NULL,
                    payload        TEXT NOT NULL,
                    undone         INTEGER NOT NULL DEFAULT 0,
                    created_at     TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_sale_items_sale ON sale_items(sale_id);
                CREATE INDEX IF NOT EXISTS idx_stock_logs_product ON stock_logs(product_id);
                CREATE INDEX IF NOT EXISTS idx_members_phone ON members(phone);
                CREATE INDEX IF NOT EXISTS idx_product_operation_logs_active
                    ON product_operation_logs(undone, id);
                """
            )
            # 老库迁移
            prod_cols = {r["name"] for r in conn.execute("PRAGMA table_info(products)").fetchall()}
            if "member_price" not in prod_cols:
                conn.execute("ALTER TABLE products ADD COLUMN member_price REAL")
            sale_cols = {r["name"] for r in conn.execute("PRAGMA table_info(sales)").fetchall()}
            if "member_id" not in sale_cols:
                conn.execute("ALTER TABLE sales ADD COLUMN member_id INTEGER")
            mem_cols = {r["name"] for r in conn.execute("PRAGMA table_info(members)").fetchall()}
            if "gender" not in mem_cols:
                conn.execute("ALTER TABLE members ADD COLUMN gender TEXT NOT NULL DEFAULT ''")
            if "points" not in mem_cols:
                conn.execute("ALTER TABLE members ADD COLUMN points INTEGER NOT NULL DEFAULT 0")
            sale_cols2 = {r["name"] for r in conn.execute("PRAGMA table_info(sales)").fetchall()}
            if "returned" not in sale_cols2:
                conn.execute("ALTER TABLE sales ADD COLUMN returned INTEGER NOT NULL DEFAULT 0")
            if "order_no" not in sale_cols2:
                conn.execute("ALTER TABLE sales ADD COLUMN order_no TEXT DEFAULT ''")
                conn.execute("CREATE INDEX IF NOT EXISTS idx_sales_order_no ON sales(order_no)")
            tables = {r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
            if "daily_counter" not in tables:
                conn.execute(
                    "CREATE TABLE daily_counter (date TEXT PRIMARY KEY, seq INTEGER NOT NULL DEFAULT 0)"
                )
    def _seed_sample(self, conn: sqlite3.Connection) -> None:
        for barcode, name, cat, cost, sell, mprice, stock, low in SAMPLE_PRODUCTS:
            conn.execute(
                "INSERT INTO products (barcode, name, category, cost_price, sell_price, member_price, stock, low_stock, created_at) "
                "VALUES (?,?,?,?,?,?,?,?,?)",
                (barcode, name, cat, cost, sell, mprice, stock, low, _now()),
            )

    # ------------------------------------------------------------------ #
    # 商品
    # ------------------------------------------------------------------ #
    @staticmethod
    def _product_snapshot(row: sqlite3.Row) -> Dict[str, Any]:
        """将商品数据库行转换为可序列化快照。"""
        return dict(row)

    def _write_product_operation(
        self, conn: sqlite3.Connection, operation_type: str, payload: Dict[str, Any]
    ) -> None:
        """在当前事务中写入商品管理撤回日志。"""
        conn.execute(
            "INSERT INTO product_operation_logs "
            "(operation_type, payload, undone, created_at) VALUES (?,?,0,?)",
            (operation_type, json.dumps(payload, ensure_ascii=False), _now()),
        )

    def add_product(
        self, name: str, category: str, cost_price: float, sell_price: float,
        stock: int, low_stock: int, barcode: Optional[str] = None,
        member_price: Optional[float] = None,
    ) -> int:
        """新增商品并写入可撤回的商品操作日志。"""
        try:
            with self._connect() as conn:
                cur = conn.execute(
                    "INSERT INTO products (barcode, name, category, cost_price, sell_price, member_price, stock, low_stock, created_at) "
                    "VALUES (?,?,?,?,?,?,?,?,?)",
                    (barcode, name, category, cost_price, sell_price, member_price, stock, low_stock, _now()),
                )
                pid = int(cur.lastrowid)
                if stock > 0:
                    conn.execute(
                        "INSERT INTO stock_logs (product_id, product_name, change_type, quantity, stock_after, note, created_at) "
                        "VALUES (?,?,?,?,?,?,?)",
                        (pid, name, "入库", stock, stock, "期初入库", _now()),
                    )
                row = conn.execute("SELECT * FROM products WHERE id=?", (pid,)).fetchone()
                self._write_product_operation(conn, "新增", {"after": self._product_snapshot(row)})
                return pid
        except sqlite3.IntegrityError as exc:
            logger.exception("新增商品失败")
            raise DatabaseError("商品条码已存在或数据不合法") from exc

    def update_product(self, pid: int, name: str, category: str,
                       cost_price: float, sell_price: float, stock: int, low_stock: int,
                       barcode: Optional[str] = None, member_price: Optional[float] = None) -> None:
        """更新商品信息、库存日志及可撤回操作日志。"""
        try:
            with self._connect() as conn:
                old = conn.execute("SELECT * FROM products WHERE id=?", (pid,)).fetchone()
                if old is None:
                    raise DatabaseError("商品不存在")
                old_snapshot = self._product_snapshot(old)
                conn.execute(
                    "UPDATE products SET barcode=?, name=?, category=?, cost_price=?, sell_price=?, "
                    "member_price=?, stock=?, low_stock=? WHERE id=?",
                    (barcode, name, category, cost_price, sell_price, member_price, stock, low_stock, pid),
                )
                delta = stock - int(old["stock"])
                if delta != 0:
                    ct = "入库" if delta > 0 else "出库"
                    conn.execute(
                        "INSERT INTO stock_logs (product_id, product_name, change_type, quantity, stock_after, note, created_at) "
                        "VALUES (?,?,?,?,?,?,?)",
                        (pid, name, ct, delta, stock, "手动调整", _now()),
                    )
                current = conn.execute("SELECT * FROM products WHERE id=?", (pid,)).fetchone()
                self._write_product_operation(
                    conn, "编辑", {"before": old_snapshot, "after": self._product_snapshot(current)}
                )
        except sqlite3.IntegrityError as exc:
            logger.exception("编辑商品失败: id=%s", pid)
            raise DatabaseError("商品条码已存在或数据不合法") from exc

    def delete_product(self, pid: int) -> None:
        """删除商品并保留完整快照以供撤回。"""
        try:
            with self._connect() as conn:
                old = conn.execute("SELECT * FROM products WHERE id=?", (pid,)).fetchone()
                if old is None:
                    raise DatabaseError("商品不存在")
                conn.execute("DELETE FROM products WHERE id=?", (pid,))
                self._write_product_operation(conn, "删除", {"before": self._product_snapshot(old)})
        except sqlite3.IntegrityError as exc:
            logger.exception("删除商品失败: id=%s", pid)
            raise DatabaseError("该商品已有销售记录，不能删除；可改名或将库存设为 0") from exc

    def batch_delete_products(self, product_ids: List[int]) -> Dict[str, int]:
        """事务性批量删除商品，并保存全部商品快照以支持撤回。"""
        unique_ids = list(dict.fromkeys(int(pid) for pid in product_ids))
        if not unique_ids:
            raise DatabaseError("未选择商品")
        deleted: List[Dict[str, Any]] = []
        skipped = 0
        try:
            with self._connect() as conn:
                for pid in unique_ids:
                    row = conn.execute("SELECT * FROM products WHERE id=?", (pid,)).fetchone()
                    if row is None:
                        skipped += 1
                        continue
                    deleted.append(self._product_snapshot(row))
                    conn.execute("DELETE FROM products WHERE id=?", (pid,))
                if deleted:
                    self._write_product_operation(conn, "批量删除", {"before": deleted})
            return {"deleted": len(deleted), "skipped": skipped}
        except sqlite3.IntegrityError as exc:
            logger.exception("批量删除商品失败")
            raise DatabaseError("所选商品已有销售记录，不能删除；可改名或将库存设为 0") from exc
        except Exception:
            logger.exception("批量删除商品发生未知异常")
            raise

    def batch_update_product_category(
        self, product_ids: List[int], target_category: str
    ) -> Dict[str, int]:
        """批量更新商品类别，返回成功数和跳过数。"""
        category = target_category.strip()
        if not category:
            raise DatabaseError("目标类别不能为空")
        unique_ids = list(dict.fromkeys(int(pid) for pid in product_ids))
        changed: List[Dict[str, Any]] = []
        skipped = 0
        try:
            with self._connect() as conn:
                for pid in unique_ids:
                    row = conn.execute("SELECT id, name, category FROM products WHERE id=?", (pid,)).fetchone()
                    if row is None or (row["category"] or "其他") == category:
                        skipped += 1
                        continue
                    changed.append({"id": int(row["id"]), "name": row["name"],
                                    "before_category": row["category"] or "其他",
                                    "after_category": category})
                    conn.execute("UPDATE products SET category=? WHERE id=?", (category, pid))
                if changed:
                    self._write_product_operation(conn, "批量类别", {"items": changed})
            return {"updated": len(changed), "skipped": skipped}
        except Exception:
            logger.exception("批量更新商品类别失败")
            raise

    def get_last_product_operation(self) -> Optional[Dict[str, Any]]:
        """返回最后一条尚未撤回的商品管理操作。"""
        with self._connect() as conn:
            row = conn.execute(
                "SELECT * FROM product_operation_logs WHERE undone=0 ORDER BY id DESC LIMIT 1"
            ).fetchone()
            if row is None:
                return None
            result = dict(row)
            result["payload"] = json.loads(result["payload"])
            return result

    def undo_last_product_operation(self) -> Dict[str, Any]:
        """事务性撤回最后一次商品管理操作。"""
        try:
            with self._connect() as conn:
                row = conn.execute(
                    "SELECT * FROM product_operation_logs WHERE undone=0 ORDER BY id DESC LIMIT 1"
                ).fetchone()
                if row is None:
                    raise DatabaseError("没有可撤回的商品管理操作")
                payload = json.loads(row["payload"])
                operation_type = row["operation_type"]
                if operation_type == "新增":
                    pid = int(payload["after"]["id"])
                    if conn.execute("SELECT 1 FROM products WHERE id=?", (pid,)).fetchone():
                        conn.execute("DELETE FROM products WHERE id=?", (pid,))
                elif operation_type == "删除":
                    self._restore_product_snapshot(conn, payload["before"], insert=True)
                elif operation_type == "批量删除":
                    for snapshot in payload.get("before", []):
                        self._restore_product_snapshot(conn, snapshot, insert=True)
                elif operation_type == "编辑":
                    self._restore_product_snapshot(conn, payload["before"], insert=False)
                elif operation_type == "批量类别":
                    for item in payload.get("items", []):
                        conn.execute("UPDATE products SET category=? WHERE id=?",
                                     (item["before_category"], int(item["id"])))
                else:
                    raise DatabaseError("无法识别的商品操作类型")
                conn.execute("UPDATE product_operation_logs SET undone=1 WHERE id=?", (row["id"],))
                return {"id": int(row["id"]), "operation_type": operation_type}
        except sqlite3.IntegrityError as exc:
            logger.exception("撤回商品操作失败")
            raise DatabaseError("撤回失败：商品已被后续业务数据引用或条码发生冲突") from exc

    @staticmethod
    def _restore_product_snapshot(
        conn: sqlite3.Connection, snapshot: Dict[str, Any], insert: bool
    ) -> None:
        """插入或覆盖恢复一个商品快照。"""
        fields = ("id", "barcode", "name", "category", "cost_price", "sell_price",
                  "member_price", "stock", "low_stock", "created_at")
        values = tuple(snapshot.get(field) for field in fields)
        if insert:
            conn.execute(
                "INSERT INTO products (id, barcode, name, category, cost_price, sell_price, member_price, stock, low_stock, created_at) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)", values
            )
        else:
            conn.execute(
                "UPDATE products SET barcode=?, name=?, category=?, cost_price=?, sell_price=?, "
                "member_price=?, stock=?, low_stock=?, created_at=? WHERE id=?",
                values[1:] + (values[0],),
            )

    def search_products(self, keyword: str = "") -> List[Dict[str, Any]]:
        with self._connect() as conn:
            if keyword.strip():
                like = f"%{keyword.strip()}%"
                rows = conn.execute(
                    "SELECT * FROM products WHERE name LIKE ? OR barcode LIKE ? OR category LIKE ? ORDER BY id DESC",
                    (like, like, like),
                ).fetchall()
                products = [dict(r) for r in rows]
                # 反向条码匹配：输入长条码但库存只存后N位时也能命中（≥3位连续相同）
                kw = keyword.strip()
                if kw.isdigit() and len(kw) >= 3:
                    all_rows = conn.execute("SELECT * FROM products").fetchall()
                    seen = {p["id"] for p in products}
                    for r in all_rows:
                        if r["id"] in seen:
                            continue
                        bc = (r["barcode"] or "").strip()
                        if not bc:
                            continue
                        if self._has_lcs(bc, kw, 3):
                            products.append(dict(r))
                return products
            else:
                rows = conn.execute("SELECT * FROM products ORDER BY id DESC").fetchall()
                return [dict(r) for r in rows]

    @staticmethod
    def _has_lcs(a: str, b: str, min_len: int) -> bool:
        """判断两个字符串是否存在至少 min_len 的连续公共子串。"""
        if len(a) < min_len or len(b) < min_len:
            return False
        for i in range(len(a)):
            for j in range(len(b)):
                if a[i] != b[j]:
                    continue
                k = 1
                while i + k < len(a) and j + k < len(b) and a[i + k] == b[j + k]:
                    k += 1
                if k >= min_len:
                    return True
        return False

    def get_product_categories(self) -> List[str]:
        """返回全量商品的去重、非空且稳定排序的类别候选。"""
        try:
            with self._connect() as conn:
                rows = conn.execute("SELECT category FROM products").fetchall()
            categories = {str(row["category"] or "").strip() for row in rows}
            categories.discard("")
            categories.add("其他")
            return sorted(categories, key=lambda value: (value.casefold(), value))
        except sqlite3.Error as exc:
            logger.exception("查询全部商品类别失败")
            raise DatabaseError("无法读取商品类别") from exc

    def get_product(self, pid: int) -> Optional[Dict[str, Any]]:
        with self._connect() as conn:
            row = conn.execute("SELECT * FROM products WHERE id=?", (pid,)).fetchone()
            return dict(row) if row else None

    def adjust_member_points(self, member_id: int, delta: int, remark: str) -> None:
        """调整会员积分，生成销售记录＋明细（sale_items 含备注）。"""
        with self._connect() as conn:
            member = conn.execute(
                "SELECT phone, name, points FROM members WHERE id=?", (member_id,)
            ).fetchone()
            if member is None:
                raise DatabaseError("会员不存在")
            new_pts = member["points"] + delta
            if new_pts < 0:
                raise DatabaseError(f"积分不足（当前 {member['points']}，变动 {delta:+d}）")
            conn.execute("UPDATE members SET points = ? WHERE id = ?", (new_pts, member_id))

            member_name = (member["name"] or "").strip() or member["phone"]
            order_no = f"积分兑换({member_name}|{delta:+d}|{remark})"

            conn.execute(
                "INSERT INTO sales (total_amount, discount, paid_amount, change_amount, member_id, created_at, order_no) "
                "VALUES (0, 0, 0, 0, ?, ?, ?)",
                (member_id, _now(), order_no),
            )

    def get_products_by_barcode(self, barcode: str) -> List[Dict[str, Any]]:
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT * FROM products WHERE barcode=?", (barcode.strip(),)
            ).fetchall()
            return [dict(r) for r in rows]

    # ------------------------------------------------------------------ #
    # 会员
    # ------------------------------------------------------------------ #
    def add_member(self, phone: str, gender: str = "", name: str = "") -> int:
        """新增会员。手机号唯一。返回会员 ID。"""
        try:
            with self._connect() as conn:
                cur = conn.execute(
                    "INSERT INTO members (phone, name, gender, created_at) VALUES (?,?,?,?)",
                    (phone.strip(), name.strip() or "", gender.strip() or "", _now()),
                )
                return int(cur.lastrowid)
        except sqlite3.IntegrityError:
            raise DatabaseError(f"手机号「{phone}」已注册，请勿重复添加") from None

    def update_member(self, mid: int, phone: str, gender: str = "", name: str = "") -> None:
        try:
            with self._connect() as conn:
                conn.execute(
                    "UPDATE members SET phone=?, name=?, gender=? WHERE id=?",
                    (phone.strip(), name.strip() or "", gender.strip() or "", mid),
                )
        except sqlite3.IntegrityError:
            raise DatabaseError(f"手机号「{phone}」已被其他会员使用") from None

    def delete_member(self, mid: int) -> None:
        with self._connect() as conn:
            conn.execute("DELETE FROM members WHERE id=?", (mid,))

    def search_members(self, keyword: str = "") -> List[Dict[str, Any]]:
        with self._connect() as conn:
            if keyword.strip():
                like = f"%{keyword.strip()}%"
                rows = conn.execute(
                    "SELECT * FROM members WHERE phone LIKE ? OR name LIKE ? ORDER BY id DESC",
                    (like, like),
                ).fetchall()
            else:
                rows = conn.execute("SELECT * FROM members ORDER BY id DESC").fetchall()
            return [dict(r) for r in rows]

    def search_members_fuzzy(self, keyword: str) -> List[Dict[str, Any]]:
        """模糊搜索: keyword >= 3 位时用 LIKE，否则前缀精确匹配。"""
        kw = keyword.strip()
        if not kw:
            return self.search_members("")
        with self._connect() as conn:
            if len(kw) >= 3:
                like = f"%{kw}%"
                rows = conn.execute(
                    "SELECT * FROM members WHERE phone LIKE ? OR name LIKE ? ORDER BY id DESC",
                    (like, like),
                ).fetchall()
            else:
                like = f"{kw}%"
                rows = conn.execute(
                    "SELECT * FROM members WHERE phone LIKE ? OR name LIKE ? ORDER BY id DESC",
                    (like, like),
                ).fetchall()
            return [dict(r) for r in rows]

    def get_member(self, mid: int) -> Optional[Dict[str, Any]]:
        with self._connect() as conn:
            row = conn.execute("SELECT * FROM members WHERE id=?", (mid,)).fetchone()
            return dict(row) if row else None

    def get_member_by_phone(self, phone: str) -> Optional[Dict[str, Any]]:
        with self._connect() as conn:
            row = conn.execute("SELECT * FROM members WHERE phone=?", (phone.strip(),)).fetchone()
            return dict(row) if row else None

    # ------------------------------------------------------------------ #
    # 销售
    # ------------------------------------------------------------------ #
    def _next_order_no(self, conn: sqlite3.Connection) -> str:
        """在事务内获取次日重置的流水单号，格式 YYYYMMDD-NNNN。"""
        today = datetime.date.today().strftime("%Y%m%d")
        conn.execute(
            "INSERT INTO daily_counter (date, seq) VALUES (?, 1) "
            "ON CONFLICT(date) DO UPDATE SET seq = seq + 1",
            (today,),
        )
        seq = conn.execute("SELECT seq FROM daily_counter WHERE date=?", (today,)).fetchone()[0]
        return f"{today}-{seq:04d}"

    def create_sale(
        self, items: List[Dict[str, Any]], discount: float = 0.0,
        paid_amount: float = 0.0, member_id: Optional[int] = None,
    ) -> Tuple[int, str, float, int]:
        """创建销售单（事务内扣库存）。返回 (sale_id, order_no, 找零金额, 积分)。"""
        total = round(sum(it["price"] * it["quantity"] for it in items), 2)
        discount = round(discount, 2)
        change = round(paid_amount - (total - discount), 2) if paid_amount else 0.0
        with self._connect() as conn:
            order_no = self._next_order_no(conn)
            cur = conn.execute(
                "INSERT INTO sales (total_amount, discount, paid_amount, change_amount, member_id, created_at, order_no) "
                "VALUES (?,?,?,?,?,?,?)",
                (total, discount, paid_amount, change, member_id, _now(), order_no),
            )
            sale_id = int(cur.lastrowid)
            for it in items:
                pid, qty = int(it["product_id"]), int(it["quantity"])
                subtotal = round(it["price"] * qty, 2)
                conn.execute(
                    "INSERT INTO sale_items (sale_id, product_id, product_name, price, quantity, subtotal) "
                    "VALUES (?,?,?,?,?,?)",
                    (sale_id, pid, it["product_name"], it["price"], qty, subtotal),
                )
                row = conn.execute(
                    "SELECT name, stock, category FROM products WHERE id=?", (pid,)
                ).fetchone()
                if row is None:
                    raise DatabaseError(f"商品「{it['product_name']}」不存在或已被删除")
                new_stock = row["stock"] - qty
                if new_stock < 0:
                    raise DatabaseError(f"商品「{row['name']}」库存不足（当前 {row['stock']}，需要 {qty}）")
                conn.execute("UPDATE products SET stock=? WHERE id=?", (new_stock, pid))
                conn.execute(
                    "INSERT INTO stock_logs (product_id, product_name, change_type, quantity, stock_after, note, created_at) "
                    "VALUES (?,?,?,?,?,?,?)",
                    (pid, row["name"], "出库", -qty, new_stock, f"销售单 #{sale_id}", _now()),
                )
                # 累计非烟类商品金额用于积分
                if row["category"] != "烟":
                    it["_non_tobacco_subtotal"] = subtotal
                else:
                    it["_non_tobacco_subtotal"] = 0.0

            # 积分计算：仅会员、仅非烟类商品，1元=1分
            points = 0
            if member_id:
                non_tobacco_total = round(sum(it.get("_non_tobacco_subtotal", 0.0) for it in items), 2)
                if non_tobacco_total > 0 and total > 0:
                    points = int((non_tobacco_total / total) * (total - discount))
                    if points > 0:
                        conn.execute(
                            "UPDATE members SET points = points + ? WHERE id = ?",
                            (points, member_id),
                        )
            return sale_id, order_no, change, points

    def get_product_sales_summary(self) -> List[Dict[str, Any]]:
        """统计所有有销售记录的商品：累计销量、累计销售额。"""
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT si.product_id, p.name, p.category, "
                "SUM(si.quantity) AS total_qty, "
                "ROUND(SUM(si.quantity * si.price), 2) AS total_revenue "
                "FROM sale_items si "
                "JOIN sales s ON s.id = si.sale_id "
                "JOIN products p ON p.id = si.product_id "
                "WHERE s.returned = 0 "
                "GROUP BY si.product_id "
                "ORDER BY total_qty DESC"
            ).fetchall()
            return [dict(r) for r in rows]

    def get_product_sale_details(self, product_id: int) -> List[Dict[str, Any]]:
        """查询某商品的所有销售明细：日期/单号/数量/单价/小计。"""
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT s.id AS sale_id, s.order_no, s.created_at, "
                "si.quantity, si.price, si.subtotal "
                "FROM sale_items si "
                "JOIN sales s ON s.id = si.sale_id "
                "WHERE si.product_id = ? AND s.returned = 0 "
                "ORDER BY s.created_at DESC",
                (product_id,),
            ).fetchall()
            return [dict(r) for r in rows]

    def search_sales(self, keyword: str = "", start: str = "", end: str = "") -> List[Dict[str, Any]]:
        conds = []
        params: List[Any] = []
        if keyword.strip():
            conds.append("(s.order_no LIKE ? OR CAST(s.id AS TEXT) LIKE ? OR si.product_name LIKE ?)")
            like = f"%{keyword.strip()}%"
            params.extend([like, like, like])
        if start:
            conds.append("s.created_at >= ?")
            params.append(start + " 00:00:00")
        if end:
            conds.append("s.created_at <= ?")
            params.append(end + " 23:59:59")
        where = (" WHERE " + " AND ".join(conds)) if conds else ""
        with self._connect() as conn:
            rows = conn.execute(
                f"SELECT s.*, (SELECT COUNT(*) FROM sale_items WHERE sale_id=s.id) AS item_count "
                f"FROM sales s {where} ORDER BY s.id DESC",
                params,
            ).fetchall()
            return [dict(r) for r in rows]

    def get_sale_items(self, sale_id: int) -> List[Dict[str, Any]]:
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT * FROM sale_items WHERE sale_id=? ORDER BY id", (sale_id,)
            ).fetchall()
            return [dict(r) for r in rows]

    def get_sale(self, sale_id: int) -> Optional[Dict[str, Any]]:
        with self._connect() as conn:
            row = conn.execute("SELECT * FROM sales WHERE id=?", (sale_id,)).fetchone()
            return dict(row) if row else None

    def get_recent_sale_ids(self) -> List[int]:
        with self._connect() as conn:
            rows = conn.execute("SELECT id FROM sales ORDER BY id DESC LIMIT 500").fetchall()
            return [r["id"] for r in rows]

    def return_sale(self, sale_id: int) -> Dict[str, Any]:
        """整单退货：恢复库存、标记已退、写日志。返回退货摘要。"""
        with self._connect() as conn:
            sale = conn.execute("SELECT * FROM sales WHERE id=?", (sale_id,)).fetchone()
            if sale is None:
                raise DatabaseError("销售记录不存在")
            if sale["returned"]:
                raise DatabaseError("该单已退货，不能重复操作")
            items = conn.execute(
                "SELECT * FROM sale_items WHERE sale_id=?", (sale_id,)
            ).fetchall()
            if not items:
                raise DatabaseError("销售明细为空")
            for it in items:
                pid = int(it["product_id"])
                product = conn.execute(
                    "SELECT name, stock FROM products WHERE id=?", (pid,)
                ).fetchone()
                if product is None:
                    logger.warning("退货时商品 id=%s 不存在，跳过库存恢复", pid)
                    continue
                new_stock = product["stock"] + int(it["quantity"])
                conn.execute("UPDATE products SET stock=? WHERE id=?", (new_stock, pid))
                conn.execute(
                    "INSERT INTO stock_logs (product_id, product_name, change_type, "
                    "quantity, stock_after, note, created_at) VALUES (?,?,?,?,?,?,?)",
                    (pid, product["name"], "退货入库", int(it["quantity"]),
                     new_stock, f"退货单 #{sale_id}", _now()),
                )
            conn.execute("UPDATE sales SET returned=1 WHERE id=?", (sale_id,))
            # 退货时扣回积分
            member_id = sale["member_id"]
            if member_id:
                total = sale["total_amount"]
                discount = sale["discount"]
                non_tobacco_total = 0.0
                for it in items:
                    pid = int(it["product_id"])
                    cat = conn.execute("SELECT category FROM products WHERE id=?", (pid,)).fetchone()
                    if cat and cat["category"] != "烟":
                        non_tobacco_total += it["subtotal"]
                if non_tobacco_total > 0 and total > 0:
                    pts = int((non_tobacco_total / total) * (total - discount))
                    if pts > 0:
                        conn.execute(
                            "UPDATE members SET points = MAX(0, points - ?) WHERE id = ?",
                            (pts, member_id),
                        )
            return {
                "sale_id": sale_id,
                "total_amount": sale["total_amount"],
                "paid_amount": sale["paid_amount"],
                "item_count": len(items),
            }

    # ------------------------------------------------------------------ #
    # 库存
    # ------------------------------------------------------------------ #
    def adjust_stock(self, product_id: int, delta: int, note: str = "") -> None:
        """入库/出库（delta 正数入库、负数出库）。自动写日志。"""
        with self._connect() as conn:
            row = conn.execute("SELECT name, stock FROM products WHERE id=?", (product_id,)).fetchone()
            if row is None:
                raise DatabaseError("商品不存在")
            new_stock = row["stock"] + delta
            if new_stock < 0:
                raise DatabaseError(f"商品「{row['name']}」库存不足（当前 {row['stock']}）")
            ct = "入库" if delta > 0 else "出库"
            conn.execute("UPDATE products SET stock=? WHERE id=?", (new_stock, product_id))
            conn.execute(
                "INSERT INTO stock_logs (product_id, product_name, change_type, quantity, stock_after, note, created_at) "
                "VALUES (?,?,?,?,?,?,?)",
                (product_id, row["name"], ct, delta, new_stock, note or "", _now()),
            )

    def adjust_stock_price(self, product_id: int, new_qty: int, new_price: float, note: str = "") -> Dict[str, Any]:
        """价格波动入库：新老库存加权平均计算新进价，写入日志。返回 {old_cost, new_cost, old_stock, new_stock}。"""
        with self._connect() as conn:
            row = conn.execute(
                "SELECT name, stock, cost_price FROM products WHERE id=?", (product_id,)
            ).fetchone()
            if row is None:
                raise DatabaseError("商品不存在")
            old_stock = row["stock"]
            old_cost = row["cost_price"]
            if new_qty <= 0:
                raise DatabaseError("入库数量必须大于0")
            new_stock = old_stock + new_qty
            new_cost = round((old_stock * old_cost + new_qty * new_price) / new_stock, 2) if new_stock > 0 else new_price
            conn.execute(
                "UPDATE products SET stock=?, cost_price=? WHERE id=?",
                (new_stock, new_cost, product_id),
            )
            conn.execute(
                "INSERT INTO stock_logs (product_id, product_name, change_type, quantity, stock_after, cost_price, note, created_at) "
                "VALUES (?,?,?,?,?,?,?,?)",
                (product_id, row["name"], "价格波动入库", new_qty, new_stock, new_cost,
                 f"旧进价={old_cost:.2f}|{note or ''}", _now()),
            )
            return {"old_cost": old_cost, "new_cost": new_cost, "old_stock": old_stock, "new_stock": new_stock}

    def get_stock_logs(self, product_id: int, limit: int = 50) -> List[Dict[str, Any]]:
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT * FROM stock_logs WHERE product_id=? ORDER BY id DESC LIMIT ?",
                (product_id, limit),
            ).fetchall()
            return [dict(r) for r in rows]

    def set_product_cost_price(self, product_id: int, cost: float) -> None:
        """设置商品进价（撤销等场景使用）。"""
        with self._connect() as conn:
            conn.execute("UPDATE products SET cost_price=? WHERE id=?", (cost, product_id))

    # ------------------------------------------------------------------ #
    # 财报分析
    # ------------------------------------------------------------------ #
    def finance_report(self, daily_days: int = 14, reorder_days: int = 7) -> Dict[str, Any]:
        """返回每日/每月/每年营收 + 畅销排行 + 库存警告 + 入库建议。
        reorder_days: 入库建议用于计算日均销量的天数。"""
        today = datetime.date.today()

        def _periods(group_expr, where_expr, params):
            with self._connect() as conn:
                rows = conn.execute(
                    f"SELECT {group_expr} AS period, "
                    "SUM(COALESCE(s.total_amount,0)-COALESCE(s.discount,0)) AS revenue, "
                    "COUNT(DISTINCT s.id) AS orders "
                    f"FROM sales s WHERE {where_expr} AND s.returned=0 GROUP BY period ORDER BY period",
                    params,
                ).fetchall()
                return [dict(r) for r in rows]

        # 成本查询（按期间）
        def _costs(period_expr, where_expr, params):
            with self._connect() as conn:
                rows = conn.execute(
                    f"SELECT {period_expr} AS period, "
                    "SUM(si.quantity * COALESCE(p.cost_price, 0)) AS cost "
                    "FROM sale_items si "
                    "JOIN sales s ON s.id = si.sale_id "
                    "JOIN products p ON p.id = si.product_id "
                    f"WHERE {where_expr} AND s.returned=0 GROUP BY period",
                    params,
                ).fetchall()
                return {r["period"]: r["cost"] for r in rows}

        # 每日（近 N 天）
        daily = _periods("date(s.created_at)", "date(s.created_at) >= ?",
                         [(today - datetime.timedelta(days=daily_days - 1)).strftime("%Y-%m-%d")])
        cost_daily = _costs("date(s.created_at)", "date(s.created_at) >= ?",
                          [(today - datetime.timedelta(days=daily_days - 1)).strftime("%Y-%m-%d")])
        daily_map = {d["period"]: d for d in daily}
        daily_result = []
        for i in range(daily_days - 1, -1, -1):
            day = (today - datetime.timedelta(days=i)).strftime("%Y-%m-%d")
            d = daily_map.get(day, {"period": day, "revenue": 0, "orders": 0})
            c = cost_daily.get(day, 0)
            d["cost"] = round(c, 2)
            d["profit"] = round(d["revenue"] - d["cost"], 2)
            d["margin"] = f"{d['profit'] / d['revenue'] * 100:.1f}%" if d["revenue"] else "—"
            daily_result.append(d)

        # 每月（近 12 个月）
        first_month = (today.replace(day=1) - datetime.timedelta(days=1)).replace(day=1)
        monthly = _periods("strftime('%Y-%m', s.created_at)",
                          "strftime('%Y-%m', s.created_at) >= ?",
                          [first_month.strftime("%Y-%m")])
        cost_monthly = _costs("strftime('%Y-%m', s.created_at)",
                            "strftime('%Y-%m', s.created_at) >= ?",
                            [first_month.strftime("%Y-%m")])
        monthly_result = []
        for i in range(11, -1, -1):
            dt = today.replace(day=1) - datetime.timedelta(days=1)
            for _ in range(i):
                dt = (dt.replace(day=1) - datetime.timedelta(days=1))
            m = dt.strftime("%Y-%m")
            found = next((x for x in monthly if x["period"] == m), None)
            d = found if found else {"period": m, "revenue": 0, "orders": 0}
            c = cost_monthly.get(m, 0)
            d["cost"] = round(c, 2)
            d["profit"] = round(d["revenue"] - d["cost"], 2)
            d["margin"] = f"{d['profit'] / d['revenue'] * 100:.1f}%" if d["revenue"] else "—"
            monthly_result.append(d)

        # 每年
        yearly = _periods("strftime('%Y', s.created_at)", "1=1", [])
        cost_yearly = _costs("strftime('%Y', s.created_at)", "1=1", [])
        yearly_result = []
        for y in yearly:
            c = cost_yearly.get(y["period"], 0)
            y["cost"] = round(c, 2)
            y["profit"] = round(y["revenue"] - y["cost"], 2)
            y["margin"] = f"{y['profit'] / y['revenue'] * 100:.1f}%" if y["revenue"] else "—"
            yearly_result.append(y)

        # 库存警告
        with self._connect() as conn:
            warns = conn.execute(
                "SELECT id, barcode, name, category, stock, low_stock FROM products WHERE stock <= low_stock ORDER BY stock"
            ).fetchall()
            warnings = [dict(r) for r in warns]

        # 入库建议（根据指定天数日均销量 × 区间天数 + 预警线 - 库存）
        with self._connect() as conn:
            reorder_rows = conn.execute(
                "SELECT p.id, p.barcode, p.name, p.category, p.stock, p.low_stock, "
                "COALESCE(sold.qty, 0) AS sold_n "
                "FROM products p "
                "LEFT JOIN ("
                "  SELECT product_id, SUM(quantity) AS qty "
                "  FROM sale_items WHERE sale_id IN ("
                "    SELECT id FROM sales WHERE date(created_at) >= ?"
                "  ) GROUP BY product_id"
                ") sold ON sold.product_id = p.id",
                [(today - datetime.timedelta(days=reorder_days)).strftime("%Y-%m-%d")],
            ).fetchall()
            reorder = []
            for r in reorder_rows:
                r = dict(r)
                daily_avg = math.ceil(r["sold_n"] / reorder_days) if r["sold_n"] and reorder_days else 0
                if daily_avg > 0:
                    suggest = max(0, daily_avg * reorder_days + r["low_stock"] - r["stock"])
                else:
                    suggest = max(0, r["low_stock"] - r["stock"])
                r["daily_avg"] = daily_avg
                r["suggest"] = suggest
                if suggest > 0:
                    reorder.append(r)

        # 汇总卡片
        def _card(rows):
            rev = round(sum(r["revenue"] for r in rows), 2)
            cost = round(sum(r["cost"] for r in rows), 2)
            profit = round(rev - cost, 2)
            margin = f"{profit / rev * 100:.1f}%" if rev else "—"
            orders = sum(r["orders"] for r in rows)
            return rev, profit, margin, cost, orders

        t_rev, t_profit, t_margin, t_cost, t_orders = _card(daily_result)
        m_rows = [r for r in monthly_result if r["revenue"] > 0]
        m_rev, m_profit, m_margin, m_cost, m_orders = _card(m_rows) if m_rows else (0, 0, "—", 0, 0)
        y_all = [r for r in yearly_result if r["revenue"] > 0]
        y_rev, y_profit, y_margin, y_cost, y_orders = _card(y_all) if y_all else (0, 0, "—", 0, 0)
        all_rows = [r for r in yearly_result if r["revenue"] > 0]
        a_rev, a_profit, a_margin, a_cost, a_orders = _card(all_rows) if all_rows else (0, 0, "—", 0, 0)

        # 库存金额 = SUM(成本 × 库存)
        with self._connect() as conn:
            inv = conn.execute(
                "SELECT SUM(stock * cost_price) AS total FROM products"
            ).fetchone()
            inventory_value = round(inv["total"] or 0, 2)

        return {
            "cards": {
                "today": (t_rev, t_profit, t_margin, t_cost, t_orders),
                "month": (m_rev, m_profit, m_margin, m_cost, m_orders),
                "year": (y_rev, y_profit, y_margin, y_cost, y_orders),
                "all": (a_rev, a_profit, a_margin, a_cost, a_orders),
            },
            "inventory_value": inventory_value,
            "daily": daily_result,
            "monthly": monthly_result,
            "yearly": yearly_result,
            "warnings": warnings,
            "reorder": reorder,
            "top_by_qty": self._top_products_sql(since_days=None, limit=10, sort_by="qty"),
            "top_by_rev": self._top_products_sql(since_days=None, limit=10, sort_by="revenue"),
        }

    @staticmethod
    def _reorder_since_date(days: int) -> str:
        """返回 N 天前的日期字符串。"""
        return (datetime.date.today() - datetime.timedelta(days=days)).strftime("%Y-%m-%d")

    def _top_products_sql(self, since_days: int = None, limit: int = 10,
                          sort_by: str = "qty") -> List[Dict[str, Any]]:
        """畅销排行：按销量或销售额排序。since_days=None 表示全部历史。"""
        where = ""
        params: list = []
        if since_days:
            where = "AND date(s.created_at) >= ?"
            params.append(self._reorder_since_date(since_days))
        order = "SUM(si.quantity) DESC" if sort_by == "qty" else "SUM(si.quantity * si.price) DESC"
        with self._connect() as conn:
            rows = conn.execute(
                f"SELECT si.product_id, p.name, p.category, "
                f"SUM(si.quantity) AS total_qty, "
                f"ROUND(SUM(si.quantity * si.price), 2) AS total_revenue "
                f"FROM sale_items si "
                f"JOIN sales s ON s.id = si.sale_id "
                f"JOIN products p ON p.id = si.product_id "
                f"WHERE s.returned = 0 {where} "
                f"GROUP BY si.product_id "
                f"ORDER BY {order} "
                f"LIMIT ?",
                params + [limit],
            ).fetchall()
            return [dict(r) for r in rows]

    def finance_range(self, start_date: str, end_date: str) -> Dict[str, Any]:
        """查询自定义日期范围的营收汇总。返回 daily + totals。"""
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT date(s.created_at) AS period, "
                "SUM(COALESCE(s.total_amount,0)-COALESCE(s.discount,0)) AS revenue, "
                "COUNT(DISTINCT s.id) AS orders "
                "FROM sales s "
                "WHERE date(s.created_at) >= ? AND date(s.created_at) <= ? AND s.returned=0 "
                "GROUP BY period ORDER BY period",
                (start_date, end_date),
            ).fetchall()
            daily = [dict(r) for r in rows]

            cost_rows = conn.execute(
                "SELECT date(s.created_at) AS period, "
                "SUM(si.quantity * COALESCE(p.cost_price, 0)) AS cost "
                "FROM sale_items si "
                "JOIN sales s ON s.id = si.sale_id "
                "JOIN products p ON p.id = si.product_id "
                "WHERE date(s.created_at) >= ? AND date(s.created_at) <= ? AND s.returned=0 "
                "GROUP BY period",
                (start_date, end_date),
            ).fetchall()
            cost_map = {r["period"]: r["cost"] for r in cost_rows}

        result = []
        for d in daily:
            c = cost_map.get(d["period"], 0)
            d["cost"] = round(c, 2)
            d["profit"] = round(d["revenue"] - d["cost"], 2)
            d["margin"] = f"{d['profit'] / d['revenue'] * 100:.1f}%" if d["revenue"] else "—"
            result.append(d)

        rev = round(sum(d["revenue"] for d in result), 2)
        cost = round(sum(d["cost"] for d in result), 2)
        profit = round(rev - cost, 2)
        margin = f"{profit / rev * 100:.1f}%" if rev else "—"
        orders = sum(d["orders"] for d in result)

        return {
            "daily": result,
            "total_revenue": rev,
            "total_cost": cost,
            "total_profit": profit,
            "total_margin": margin,
            "total_orders": orders,
        }

    # ------------------------------------------------------------------ #
    # 备份
    # ------------------------------------------------------------------ #
    def backup(self, backup_dir: str) -> str:
        """使用 SQLite Online Backup API 创建包含 WAL 数据的一致性数据库备份。"""
        os.makedirs(backup_dir, exist_ok=True)
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = os.path.join(backup_dir, f"cashier_backup_{ts}.db")
        try:
            with self._connect() as source, sqlite3.connect(dest) as target:
                source.backup(target)
            with sqlite3.connect(dest) as check:
                result = check.execute("PRAGMA integrity_check").fetchone()[0]
            if result != "ok":
                raise DatabaseError("备份数据库完整性校验失败")
            return dest
        except Exception as exc:
            try:
                if os.path.isfile(dest):
                    os.remove(dest)
            except OSError:
                logger.exception("删除无效备份文件失败")
            logger.exception("数据库一致性备份失败")
            if isinstance(exc, DatabaseError):
                raise
            raise DatabaseError("数据库备份失败") from exc

    def reset_all_data(self) -> None:
        """以单一事务清空全部业务数据并保留数据库结构。"""
        try:
            with self._connect() as conn:
                conn.execute("BEGIN IMMEDIATE")
                conn.execute("DELETE FROM sale_items")
                conn.execute("DELETE FROM sales")
                conn.execute("DELETE FROM stock_logs")
                conn.execute("DELETE FROM product_operation_logs")
                conn.execute("DELETE FROM members")
                conn.execute("DELETE FROM products")
                conn.execute(
                    "DELETE FROM sqlite_sequence WHERE name IN "
                    "('sale_items','sales','stock_logs','product_operation_logs','members','products')"
                )
        except Exception as exc:
            logger.exception("重置全部业务数据失败")
            raise DatabaseError("重置软件数据失败") from exc

    # ------------------------------------------------------------------ #
    # 导出
    # ------------------------------------------------------------------ #
    def _member_phone_map(self) -> Dict[int, str]:
        return {m["id"]: m["phone"] for m in self.search_members()}

    def export_products_csv(self, path: str) -> int:
        rows = self.search_products()
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "条码", "名称", "类别", "进价", "售价", "会员价", "库存", "预警线", "创建时间"])
            for r in rows:
                writer.writerow(
                    [r["id"], r["barcode"], r["name"], r["category"],
                     r["cost_price"], r["sell_price"], r.get("member_price") or "",
                     r["stock"], r["low_stock"], r["created_at"]]
                )
        return len(rows)

    def export_sales_csv(self, path: str) -> int:
        member_map = self._member_phone_map()
        sales = self.search_sales()
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["销售单号", "会员", "状态", "商品", "数量", "单价", "小计", "总金额", "优惠", "实收", "找零", "时间"])
            for s in sales:
                items = self.get_sale_items(s["id"])
                member = member_map.get(s["member_id"]) or ""
                status = "已退货" if s.get("returned") else ""
                for it in items:
                    writer.writerow(
                        [s.get("order_no") or s["id"], member, status, it["product_name"], it["quantity"],
                         it["price"], it["subtotal"], s["total_amount"],
                         s["discount"], s["paid_amount"], s["change_amount"], s["created_at"]]
                    )
        return len(sales)

    def export_products_excel(self, path: str) -> int:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "商品清单"
        ws.append(["ID", "条码", "名称", "类别", "进价", "售价", "会员价", "库存", "预警线", "创建时间"])
        for r in self.search_products():
            ws.append([r["id"], r["barcode"], r["name"], r["category"],
                       r["cost_price"], r["sell_price"], r.get("member_price") or "",
                       r["stock"], r["low_stock"], r["created_at"]])
        self._style_sheet(ws, [5, 6, 7, 8])
        wb.save(path)
        return len(self.search_products())

    def export_sales_excel(self, path: str) -> int:
        from openpyxl import Workbook
        member_map = self._member_phone_map()
        wb = Workbook()
        ws = wb.active
        ws.title = "销售记录"
        ws.append(["销售单号", "会员", "状态", "商品", "数量", "单价", "小计", "总金额", "优惠", "实收", "找零", "时间"])
        for s in self.search_sales():
            items = self.get_sale_items(s["id"])
            member = member_map.get(s["member_id"]) or ""
            status = "已退货" if s.get("returned") else ""
            for it in items:
                ws.append([s.get("order_no") or s["id"], member, status, it["product_name"], it["quantity"],
                           it["price"], it["subtotal"], s["total_amount"],
                           s["discount"], s["paid_amount"], s["change_amount"], s["created_at"]])
        self._style_sheet(ws, [5, 6, 7, 8])
        wb.save(path)
        return len(self.search_sales())

    # --- 会员导出 ---
    def export_members_csv(self, path: str) -> int:
        rows = self.search_members()
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "手机号", "姓名", "性别", "会员等级", "积分", "创建时间"])
            for r in rows:
                writer.writerow([r["id"], r["phone"], r.get("name") or "",
                                 r.get("gender") or "", r.get("level") or "",
                                 r.get("points") or 0, r.get("created_at") or ""])
        return len(rows)

    def export_members_excel(self, path: str) -> int:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "会员清单"
        ws.append(["ID", "手机号", "姓名", "性别", "会员等级", "积分", "创建时间"])
        for r in self.search_members():
            ws.append([r["id"], r["phone"], r.get("name") or "",
                       r.get("gender") or "", r.get("level") or "",
                       r.get("points") or 0, r.get("created_at") or ""])
        self._style_sheet(ws, [])
        wb.save(path)
        return len(self.search_members())

    # --- 库存导出 ---
    def export_stock_csv(self, path: str) -> int:
        rows = self.search_products()
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "条码", "名称", "类别", "进价", "当前库存", "预警线"])
            for r in rows:
                writer.writerow([r["id"], r["barcode"], r["name"], r["category"],
                                 r.get("cost_price") or 0, r["stock"], r["low_stock"]])
        return len(rows)

    @staticmethod
    def _clean_note(note: str) -> str:
        """去除价格波动内部标记前缀。"""
        if note and note.startswith("旧进价=") and "|" in note:
            return note[note.index("|") + 1:]
        return note or ""

    def export_stock_logs_csv(self, path: str) -> int:
        """导出全部库存日志为 CSV。"""
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT sl.*, p.barcode FROM stock_logs sl LEFT JOIN products p ON sl.product_id=p.id "
                "ORDER BY sl.id DESC"
            ).fetchall()
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "商品ID", "条码", "商品名", "操作类型", "数量", "库存后", "进价", "备注", "时间"])
            for r in rows:
                cost_str = f"{r['cost_price']:.2f}" if r["cost_price"] is not None else ""
                writer.writerow([r["id"], r["product_id"], r["barcode"] or "", r["product_name"],
                                 r["change_type"], r["quantity"], r["stock_after"],
                                 cost_str, self._clean_note(r["note"]), r["created_at"]])
        return len(rows)

    def export_stock_logs_excel(self, path: str) -> int:
        """导出全部库存日志为 Excel。"""
        from openpyxl import Workbook
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT sl.*, p.barcode FROM stock_logs sl LEFT JOIN products p ON sl.product_id=p.id "
                "ORDER BY sl.id DESC"
            ).fetchall()
        wb = Workbook()
        ws = wb.active
        ws.title = "库存日志"
        ws.append(["ID", "商品ID", "条码", "商品名", "操作类型", "数量", "库存后", "进价", "备注", "时间"])
        for r in rows:
            ws.append([r["id"], r["product_id"], r["barcode"] or "", r["product_name"],
                       r["change_type"], r["quantity"], r["stock_after"],
                       r["cost_price"], self._clean_note(r["note"]), r["created_at"]])
        self._style_sheet(ws, [])
        wb.save(path)
        return len(rows)

    def get_all_stock_logs(self) -> List[Dict[str, Any]]:
        """获取全部库存日志（含条码）。"""
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT sl.*, p.barcode FROM stock_logs sl LEFT JOIN products p ON sl.product_id=p.id "
                "ORDER BY sl.id DESC"
            ).fetchall()
            return [dict(r) for r in rows]

    # --- 一键导出全部 ---
    def export_all_excel(self, path: str) -> int:
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)
        member_map = self._member_phone_map()

        ws1 = wb.create_sheet("商品清单")
        ws1.append(["ID", "条码", "名称", "类别", "进价", "售价", "会员价", "库存", "预警线", "创建时间"])
        for r in self.search_products():
            ws1.append([r["id"], r["barcode"], r["name"], r["category"],
                        r["cost_price"], r["sell_price"], r.get("member_price") or "",
                        r["stock"], r["low_stock"], r["created_at"]])
        self._style_sheet(ws1, [5, 6, 7, 8])

        ws2 = wb.create_sheet("会员清单")
        ws2.append(["ID", "手机号", "姓名", "性别", "会员等级", "积分", "创建时间"])
        for r in self.search_members():
            ws2.append([r["id"], r["phone"], r.get("name") or "",
                        r.get("gender") or "", r.get("level") or "",
                        r.get("points") or 0, r.get("created_at") or ""])
        self._style_sheet(ws2, [])

        ws3 = wb.create_sheet("销售记录")
        ws3.append(["销售单号", "会员", "商品", "数量", "单价", "小计", "总金额", "优惠", "实收", "找零", "时间"])
        for s in self.search_sales():
            items = self.get_sale_items(s["id"])
            member = member_map.get(s["member_id"]) or ""
            for it in items:
                ws3.append([s["id"], member, it["product_name"], it["quantity"],
                            it["price"], it["subtotal"], s["total_amount"],
                            s["discount"], s["paid_amount"], s["change_amount"], s["created_at"]])
        self._style_sheet(ws3, [5, 6, 7, 8])

        ws4 = wb.create_sheet("库存快照")
        ws4.append(["ID", "条码", "名称", "类别", "进价", "当前库存", "预警线"])
        for r in self.search_products():
            ws4.append([r["id"], r["barcode"], r["name"], r["category"],
                        r.get("cost_price") or 0, r["stock"], r["low_stock"]])
        self._style_sheet(ws4, [4, 5])

        ws5 = wb.create_sheet("库存日志")
        ws5.append(["ID", "商品ID", "条码", "商品名", "操作类型", "数量", "库存后", "进价", "备注", "时间"])
        with self._connect() as conn:
            log_rows = conn.execute(
                "SELECT sl.*, p.barcode FROM stock_logs sl LEFT JOIN products p ON sl.product_id=p.id "
                "ORDER BY sl.id DESC"
            ).fetchall()
        for r in log_rows:
            ws5.append([r["id"], r["product_id"], r["barcode"] or "", r["product_name"],
                        r["change_type"], r["quantity"], r["stock_after"],
                        r["cost_price"], self._clean_note(r["note"]), r["created_at"]])
        self._style_sheet(ws5, [])

        wb.save(path)
        total = len(self.search_products()) + len(self.search_members()) + len(self.search_sales()) + len(log_rows)
        return total

    # --- 销售导入模板 ---
    def write_sales_template(self, path: str) -> None:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "销售导入"
        ws.append(["销售单号", "商品条码", "数量", "单价", "备注"])
        ws.append(["1001", "6900000000001", "2", "19.90", "线下补录"])
        self._style_sheet(ws, [3])
        ws2 = wb.create_sheet("格式说明")
        for r in [
            ["列名", "是否必填", "说明"],
            ["销售单号", "是", "销售单号，相同单号归为同一笔（整数）"],
            ["商品条码", "是", "对应商品的条形码"],
            ["数量", "是", "销售数量（正整数）"],
            ["单价", "是", "销售单价"],
            ["备注", "否", "备注信息"],
        ]:
            ws2.append(r)
        self._style_sheet(ws2, [])
        wb.save(path)

    # --- 销售导入解析 ---
    def parse_sales_import(self, path: str) -> List[Dict[str, Any]]:
        header_map = {
            "销售单号": "sale_id", "单号": "sale_id", "sale_id": "sale_id",
            "商品条码": "barcode", "barcode": "barcode", "条码": "barcode",
            "数量": "quantity", "quantity": "quantity", "qty": "quantity",
            "单价": "price", "price": "price",
            "备注": "note", "note": "note",
        }
        out = []
        for item in self.read_table_file(path):
            row = {}
            for k, v in item.items():
                key = header_map.get((k or "").strip().lower()) if k else None
                if key:
                    row[key] = v
            if row:
                out.append(row)
        return out

    # --- 销售导入执行 ---
    def import_sales(self, rows: List[Dict[str, Any]]) -> Dict[str, int]:
        groups: Dict[int, List[Dict[str, Any]]] = {}
        for row in rows:
            try:
                sid = int(float(row.get("sale_id") or 0))
            except (TypeError, ValueError):
                continue
            if sid <= 0:
                continue
            groups.setdefault(sid, []).append(row)

        inserted = skipped = 0
        for sid, items in groups.items():
            try:
                with self._connect() as conn:
                    existing = conn.execute("SELECT id FROM sales WHERE id=?", (sid,)).fetchone()
                    if existing:
                        skipped += len(items)
                        continue
                    total = 0.0
                    sale_items = []
                    for row in items:
                        barcode = str(row.get("barcode") or "").strip()
                        if not barcode:
                            continue
                        prod = conn.execute("SELECT id, name, sell_price, stock FROM products WHERE barcode=?",
                                            (barcode,)).fetchone()
                        if not prod:
                            continue
                        try:
                            qty = int(float(row.get("quantity") or 1))
                        except (TypeError, ValueError):
                            qty = 1
                        try:
                            price = float(row.get("price") or prod["sell_price"])
                        except (TypeError, ValueError):
                            price = prod["sell_price"]
                        subtotal = price * qty
                        total += subtotal
                        sale_items.append((prod["id"], prod["name"], price, qty, subtotal))
                    if not sale_items:
                        skipped += len(items)
                        continue
                    conn.execute(
                        "INSERT INTO sales (id, total_amount, discount, paid_amount, change_amount, member_id, created_at) "
                        "VALUES (?,?,0,?,0,NULL,?)",
                        (sid, total, total, _now()))
                    for pid, pname, pr, qty, sub in sale_items:
                        conn.execute(
                            "INSERT INTO sale_items (sale_id, product_id, product_name, price, quantity, subtotal) "
                            "VALUES (?,?,?,?,?,?)",
                            (sid, pid, pname, pr, qty, sub))
                        conn.execute("UPDATE products SET stock=stock-? WHERE id=?", (qty, pid))
                    inserted += 1
            except Exception:
                skipped += len(items)
        return {"inserted": inserted, "skipped": skipped}

    # --- 一键全部导入模板 ---
    def write_all_template(self, path: str) -> None:
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)

        ws1 = wb.create_sheet("商品导入")
        ws1.append(["条码", "名称", "类别", "进价", "售价", "会员价", "库存", "预警线"])
        ws1.append(["6900000000001", "示例白酒", "白酒", "50.00", "88.00", "78.00", "100", "10"])
        self._style_sheet(ws1, [4, 5, 6, 7, 8])
        ws1a = wb.create_sheet("商品导入-说明")
        for r in [
            ["列名", "是否必填", "说明"],
            ["条码", "是", "唯一识别码"],
            ["名称", "是", "商品名称"],
            ["类别", "否"],
            ["进价", "否"],
            ["售价", "否"],
            ["会员价", "否", "不填则不启用会员价"],
            ["库存", "否", "默认 0"],
            ["预警线", "否", "默认 0"],
        ]:
            ws1a.append(r)
        self._style_sheet(ws1a, [])

        ws2 = wb.create_sheet("会员导入")
        ws2.append(["手机号", "姓名", "性别"])
        ws2.append(["13800000000", "张三", "男"])
        self._style_sheet(ws2, [])
        ws2a = wb.create_sheet("会员导入-说明")
        for r in [
            ["列名", "是否必填", "说明"],
            ["手机号", "是", "11位数字，已存在则跳过"],
            ["姓名", "否"],
            ["性别", "否", "男/女"],
        ]:
            ws2a.append(r)
        self._style_sheet(ws2a, [])

        ws3 = wb.create_sheet("库存调整")
        ws3.append(["条码", "名称", "新库存", "备注"])
        self._style_sheet(ws3, [3])
        ws3a = wb.create_sheet("库存调整-说明")
        for r in [
            ["列名", "是否必填", "说明"],
            ["条码", "二选一", "优先按条码匹配"],
            ["名称", "二选一"],
            ["新库存", "是", "调整后库存（>=0 整数）"],
            ["备注", "否"],
        ]:
            ws3a.append(r)
        self._style_sheet(ws3a, [])

        ws4 = wb.create_sheet("销售补录")
        ws4.append(["销售单号", "商品条码", "数量", "单价", "备注"])
        self._style_sheet(ws4, [3])
        ws4a = wb.create_sheet("销售补录-说明")
        for r in [
            ["列名", "是否必填", "说明"],
            ["销售单号", "是", "相同单号归为同一笔（整数）"],
            ["商品条码", "是"],
            ["数量", "是", "正整数"],
            ["单价", "是"],
            ["备注", "否"],
        ]:
            ws4a.append(r)
        self._style_sheet(ws4a, [])

        wb.save(path)

    def daily_backup_zip(self, daily_root: str) -> str:
        """打包核心数据为 ZIP：商品.xlsx + 销售.xlsx + 会员.xlsx + cashier.db。"""
        import zipfile
        day = datetime.datetime.now().strftime("%Y-%m-%d")
        day_dir = os.path.join(daily_root, day)
        os.makedirs(day_dir, exist_ok=True)

        products_xlsx = os.path.join(day_dir, "商品.xlsx")
        sales_xlsx = os.path.join(day_dir, "销售.xlsx")
        members_xlsx = os.path.join(day_dir, "会员.xlsx")

        self.export_products_excel(products_xlsx)
        self.export_sales_excel(sales_xlsx)
        self.export_members_excel(members_xlsx)
        stock_logs_xlsx = os.path.join(day_dir, "库存日志.xlsx")
        self.export_stock_logs_excel(stock_logs_xlsx)

        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        cnt = 1
        while True:
            suffix = f"_({cnt})" if cnt > 1 else ""
            zip_path = os.path.join(day_dir, f"{ts}{suffix}.zip")
            if not os.path.exists(zip_path):
                break
            cnt += 1

        db_snapshot = self.backup(day_dir)
        try:
            with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.write(db_snapshot, "cashier.db")
                zf.write(products_xlsx, "商品.xlsx")
                zf.write(sales_xlsx, "销售.xlsx")
                zf.write(members_xlsx, "会员.xlsx")
                zf.write(stock_logs_xlsx, "库存日志.xlsx")
        finally:
            try:
                os.remove(db_snapshot)
            except OSError:
                logger.exception("删除每日备份临时数据库失败")
        # 验证备份非空
        try:
            if os.path.getsize(zip_path) == 0:
                os.remove(zip_path)
                raise DatabaseError("备份文件大小为 0，已删除空文件")
        except OSError:
            pass
        return zip_path

    @staticmethod
    def cleanup_old_backups(daily_root: str) -> None:
        """清理3天前的备份：每天只保留最新1个有效ZIP。当日/昨/前天不动。"""
        if not os.path.isdir(daily_root):
            return
        today = datetime.date.today()
        cutoff = (today - datetime.timedelta(days=3)).strftime("%Y-%m-%d")

        for entry in sorted(os.listdir(daily_root)):
            dir_path = os.path.join(daily_root, entry)
            if not os.path.isdir(dir_path):
                continue
            if len(entry) != 10 or entry[4] != "-" or entry[7] != "-":
                continue

            if entry >= cutoff:
                continue  # 当日/昨/前天：不动

            # 3天前及更早：只保留最新1个有效 ZIP，其余删除
            zips = sorted(
                [f for f in os.listdir(dir_path) if f.endswith(".zip")],
                reverse=True,
            )
            if not zips:
                continue
            kept = None
            for zf in zips:
                zpath = os.path.join(dir_path, zf)
                try:
                    if os.path.getsize(zpath) > 0:
                        kept = zf
                        break
                    else:
                        os.remove(zpath)
                        logger.warning("删除空备份: %s", zpath)
                except OSError:
                    logger.exception("检查备份文件失败: %s", zpath)
            for zf in zips:
                if zf != kept:
                    try:
                        os.remove(os.path.join(dir_path, zf))
                    except OSError:
                        logger.exception("清理旧备份失败: %s", zf)

    # ------------------------------------------------------------------ #
    # 导入
    # ------------------------------------------------------------------ #
    @staticmethod
    def read_table_file(path: str) -> List[Dict[str, str]]:
        ext = os.path.splitext(path)[1].lower()
        if ext == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                ws = wb.active
                raw = list(ws.iter_rows(values_only=True))
            finally:
                wb.close()
        elif ext == ".csv":
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                raw = list(csv.reader(f))
        else:
            raise DatabaseError(f"不支持的文件格式: {ext}")
        if not raw:
            return []
        headers = [str(h or "").strip() for h in raw[0]]
        out = []
        for row in raw[1:]:
            d: Dict[str, str] = {}
            for i, cell in enumerate(row):
                if i < len(headers) and headers[i]:
                    d[headers[i]] = str(cell or "").strip()
            if d:
                out.append(d)
        return out

    @staticmethod
    def _style_sheet(ws, numeric_cols: List[int]) -> None:
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError:
            return
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for col in numeric_cols:
            col_letter = openpyxl.utils.get_column_letter(col)
            for cell in ws[col_letter][1:]:
                try:
                    cell.number_format = "0.00"
                except Exception:
                    pass

    def parse_product_import(self, path: str) -> List[Dict[str, Any]]:
        header_map = {
            "条码": "barcode", "barcode": "barcode",
            "名称": "name", "name": "name", "商品名称": "name",
            "类别": "category", "category": "category",
            "进价": "cost_price", "cost_price": "cost_price",
            "售价": "sell_price", "sell_price": "sell_price",
            "会员价": "member_price", "member_price": "member_price",
            "初始库存": "stock", "库存": "stock", "库存数量": "stock", "当前库存": "stock", "stock": "stock",
            "预警线": "low_stock", "low_stock": "low_stock",
        }
        out: List[Dict[str, Any]] = []
        for item in self.read_table_file(path):
            row: Dict[str, Any] = {}
            for k, v in item.items():
                key = header_map.get(k.strip().lower()) if k else None
                if key:
                    row[key] = v
            out.append(row)
        return out

    def import_products(self, rows: List[Dict[str, Any]]) -> Dict[str, int]:
        inserted = updated = skipped = 0
        for row in rows:
            name = str(row.get("name") or "").strip()
            if not name:
                skipped += 1
                continue
            barcode = str(row.get("barcode") or "").strip() or None
            category = str(row.get("category") or "其他").strip() or "其他"

            def _safe_float(key: str, default: float = 0.0) -> float:
                try:
                    return float(row.get(key) or default)
                except (TypeError, ValueError):
                    return default

            def _safe_int(key: str, default: int = 0) -> int:
                try:
                    return int(float(row.get(key) or default))
                except (TypeError, ValueError):
                    return default

            cost = _safe_float("cost_price")
            sell = _safe_float("sell_price")
            mp_raw = row.get("member_price")
            try:
                mp = float(mp_raw) if str(mp_raw or "").strip() else None
            except (TypeError, ValueError):
                mp = None
            stock = _safe_int("stock")
            low = _safe_int("low_stock", 10)

            with self._connect() as conn:
                exist = None
                if barcode:
                    exist = conn.execute(
                        "SELECT id, name FROM products WHERE barcode=? AND name=?",
                        (barcode, name),
                    ).fetchone()
                if exist:
                    conn.execute(
                        "UPDATE products SET name=?, category=?, cost_price=?, sell_price=?, "
                        "member_price=COALESCE(?, member_price), stock=?, low_stock=? WHERE id=?",
                        (name, category, cost, sell, mp, stock, low, exist["id"]),
                    )
                    updated += 1
                else:
                    conn.execute(
                        "INSERT INTO products (barcode, name, category, cost_price, sell_price, "
                        "member_price, stock, low_stock, created_at) "
                        "VALUES (?,?,?,?,?,?,?,?,?)",
                        (barcode, name, category, cost, sell, mp, stock, low, _now()),
                    )
                    inserted += 1
        return {"inserted": inserted, "updated": updated, "skipped": skipped}

    def write_product_template(self, path: str) -> None:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "商品导入"
        ws.append(["条码", "名称", "类别", "进价", "售价", "会员价", "初始库存", "预警线"])
        ws.append(["6900000000001", "示例商品（导入前请删除此行）", "白酒", "100", "150", "140", "10", "5"])
        self._style_sheet(ws, [])
        ws2 = wb.create_sheet("格式说明")
        for r in [
            ["列名", "是否必填", "说明"],
            ["条码", "否", "留空=无条码商品；与库中条码重复时更新该商品信息"],
            ["名称", "是", "商品名称，必填"],
            ["类别", "否", "留空默认「其他」"],
            ["进价", "否", "进货成本价（数字）"],
            ["售价", "否", "零售价（数字）"],
            ["会员价", "否", "会员折扣价；留空=无会员折扣"],
            ["初始库存", "否", "新增或已存在商品均会写入库存数量；留空默认 0"],
            ["预警线", "否", "库存预警阈值，默认 10"],
        ]:
            ws2.append(r)
        self._style_sheet(ws2, [])
        wb.save(path)

    def parse_member_import(self, path: str) -> List[Dict[str, Any]]:
        header_map = {
            "手机号": "phone", "phone": "phone", "手机": "phone", "会员手机号": "phone",
            "姓名": "name", "name": "name", "会员姓名": "name",
            "性别": "gender", "gender": "gender", "sex": "gender",
        }
        out: List[Dict[str, Any]] = []
        for item in self.read_table_file(path):
            row: Dict[str, Any] = {}
            for k, v in item.items():
                key = header_map.get(k.strip().lower()) if k else None
                if key:
                    row[key] = v
            out.append(row)
        return out

    def import_members(self, rows: List[Dict[str, Any]]) -> Dict[str, int]:
        inserted = skipped = 0
        for row in rows:
            phone = str(row.get("phone") or "").strip()
            name = str(row.get("name") or "").strip()
            gender = str(row.get("gender") or "").strip()
            if not phone or not phone.isdigit():
                skipped += 1
                continue
            try:
                with self._connect() as conn:
                    if conn.execute("SELECT 1 FROM members WHERE phone=?", (phone,)).fetchone():
                        skipped += 1
                        continue
                    conn.execute("INSERT INTO members (phone, name, gender, created_at) VALUES (?,?,?,?)",
                                 (phone, name, gender, _now()))
                    inserted += 1
            except sqlite3.IntegrityError:
                skipped += 1
        return {"inserted": inserted, "skipped": skipped}

    def write_member_template(self, path: str) -> None:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "会员导入"
        ws.append(["手机号", "姓名（可选）", "性别（男/女/空）"])
        ws.append(["13800000000", "张三", "男"])
        ws.append(["13900000000", "", "女"])
        self._style_sheet(ws, [])
        ws2 = wb.create_sheet("格式说明")
        for r in [
            ["列名", "是否必填", "说明"],
            ["手机号", "是", "11 位数字手机号，库中已存在则跳过"],
            ["姓名", "否", "会员姓名，可不填"],
            ["性别", "否", "填「男」或「女」，可不填"],
        ]:
            ws2.append(r)
        self._style_sheet(ws2, [])
        wb.save(path)

    def write_stock_template(self, path: str) -> None:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "库存调整"
        ws.append(["条码", "名称", "新库存", "备注"])
        ws.append(["6900000000001", "示例商品（导入前请删除此行）", "50", "盘点入库"])
        self._style_sheet(ws, [])
        ws2 = wb.create_sheet("格式说明")
        for r in [
            ["列名", "是否必填", "说明"],
            ["条码", "二选一", "商品条码（优先按条码匹配）"],
            ["名称", "二选一", "商品名称（无条码时按名称模糊匹配）"],
            ["新库存", "是", "调整后的库存数量（整数 >=0）"],
            ["备注", "否", "库存变动备注"],
        ]:
            ws2.append(r)
        self._style_sheet(ws2, [])
        wb.save(path)

    def parse_stock_import(self, path: str) -> List[Dict[str, Any]]:
        header_map = {
            "条码": "barcode", "barcode": "barcode", "条形码": "barcode",
            "名称": "name", "name": "name", "商品名称": "name", "商品": "name",
            "新库存": "stock", "stock": "stock", "库存": "stock", "库存数量": "stock",
            "备注": "note", "note": "note", "说明": "note",
        }
        out: List[Dict[str, Any]] = []
        for item in self.read_table_file(path):
            row: Dict[str, Any] = {}
            for k, v in item.items():
                key = header_map.get(k.strip().lower()) if k else None
                if key:
                    row[key] = v
            out.append(row)
        return out

    def import_stock(self, rows: List[Dict[str, Any]]) -> Dict[str, int]:
        updated = skipped = 0
        for row in rows:
            barcode = str(row.get("barcode") or "").strip() or None
            name = str(row.get("name") or "").strip()
            if not barcode and not name:
                skipped += 1
                continue
            try:
                stock_new = int(float(row.get("stock") or 0))
                if stock_new < 0:
                    skipped += 1
                    continue
            except (TypeError, ValueError):
                skipped += 1
                continue
            note = str(row.get("note") or "").strip() or "Excel批量导入"
            try:
                with self._connect() as conn:
                    prod = None
                    if barcode:
                        prod = conn.execute("SELECT id, name, stock FROM products WHERE barcode=?", (barcode,)).fetchone()
                    if prod is None and name:
                        prod = conn.execute("SELECT id, name, stock FROM products WHERE name LIKE ?", (f"%{name}%",)).fetchone()
                    if prod is None:
                        skipped += 1
                        continue
                    old_stock = int(prod["stock"])
                    delta = stock_new - old_stock
                    if delta == 0:
                        skipped += 1
                        continue
                    ct = "入库" if delta > 0 else "出库"
                    conn.execute("UPDATE products SET stock=? WHERE id=?", (stock_new, prod["id"]))
                    conn.execute(
                        "INSERT INTO stock_logs (product_id, product_name, change_type, quantity, stock_after, note, created_at) "
                        "VALUES (?,?,?,?,?,?,?)",
                        (prod["id"], prod["name"], ct, delta, stock_new, note, _now()),
                    )
                    updated += 1
            except Exception:
                skipped += 1
        return {"updated": updated, "skipped": skipped}

    # ------------------------------------------------------------------ #
    # 恢复
    # ------------------------------------------------------------------ #
    def prepare_restore(self, source_path: str) -> str:
        """校验并准备恢复：解压到 staging 目录，写 .bat 换文件脚本。返回 .bat 路径。"""
        import zipfile, tempfile

        if not os.path.isfile(source_path):
            raise DatabaseError(f"备份文件不存在: {source_path}")
        if not source_path.lower().endswith(".zip"):
            raise DatabaseError("仅支持 .zip 备份包")

        # 临时解压校验
        tmpdir = tempfile.mkdtemp()
        try:
            with zipfile.ZipFile(source_path, "r") as zf:
                for name in zf.namelist():
                    if ".." in name or name.startswith("/") or name.startswith("\\"):
                        raise DatabaseError("备份包包含不安全路径")
                zf.extractall(tmpdir)
        except DatabaseError:
            shutil.rmtree(tmpdir, ignore_errors=True)
            raise
        except Exception:
            shutil.rmtree(tmpdir, ignore_errors=True)
            raise DatabaseError("ZIP 解压失败，文件可能已损坏") from None

        db_in_tmp = os.path.join(tmpdir, "cashier.db")
        if not os.path.isfile(db_in_tmp):
            shutil.rmtree(tmpdir, ignore_errors=True)
            raise DatabaseError("ZIP 包中未找到 cashier.db")

        # 校验待恢复数据库
        try:
            check = sqlite3.connect(db_in_tmp)
            required_tables = {"products", "sales", "sale_items", "members", "stock_logs"}
            actual = {r[0] for r in check.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
            missing = required_tables - actual
            if missing:
                check.close()
                shutil.rmtree(tmpdir, ignore_errors=True)
                raise DatabaseError(f"数据库缺少必要表: {missing}")
            if check.execute("PRAGMA integrity_check").fetchone()[0] != "ok":
                check.close()
                shutil.rmtree(tmpdir, ignore_errors=True)
                raise DatabaseError("数据库完整性校验失败")
            check.close()
        except DatabaseError:
            raise
        except Exception:
            shutil.rmtree(tmpdir, ignore_errors=True)
            raise DatabaseError("数据库无法打开") from None

        # 备份当前数据到备份文件夹
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        prev_name = f"恢复前备份_{ts}.zip"
        prev_path = os.path.join(self.backup_dir, prev_name)
        try:
            current_db_snapshot = self.backup(self.backup_dir)
            import zipfile as _zf
            with _zf.ZipFile(prev_path, "w", _zf.ZIP_DEFLATED) as zf:
                zf.write(current_db_snapshot, "cashier.db")
                # 也附带当前数据的 Excel 快照
                px = os.path.join(self.backup_dir, f"_prev_商品_{ts}.xlsx")
                sx = os.path.join(self.backup_dir, f"_prev_销售_{ts}.xlsx")
                mx = os.path.join(self.backup_dir, f"_prev_会员_{ts}.xlsx")
                try:
                    self.export_products_excel(px)
                    zf.write(px, "商品.xlsx")
                    self.export_sales_excel(sx)
                    zf.write(sx, "销售.xlsx")
                    self.export_members_excel(mx)
                    zf.write(mx, "会员.xlsx")
                finally:
                    for f in (px, sx, mx):
                        try:
                            os.remove(f)
                        except OSError:
                            pass
                try:
                    os.remove(current_db_snapshot)
                except OSError:
                    pass
        except Exception:
            shutil.rmtree(tmpdir, ignore_errors=True)
            raise DatabaseError("备份当前数据失败，中止恢复") from None

        # 搬到 staging
        data_dir = os.path.dirname(self.db_path)
        staging = os.path.join(data_dir, "_restore_staging")
        if os.path.exists(staging):
            shutil.rmtree(staging)
        shutil.move(tmpdir, staging)

        # 写 .bat 脚本（程序退出后由它接管覆盖）
        bat = os.path.join(data_dir, "_restore.bat")
        with open(bat, "w", encoding="gbk") as f:
            f.write(f"""@echo off
timeout /t 2 /nobreak >nul
move /Y "{staging}\\cashier.db" "{self.db_path}"
if exist "{staging}\\商品.xlsx" move /Y "{staging}\\商品.xlsx" "{data_dir}\\"
if exist "{staging}\\销售.xlsx" move /Y "{staging}\\销售.xlsx" "{data_dir}\\"
if exist "{staging}\\会员.xlsx" move /Y "{staging}\\会员.xlsx" "{data_dir}\\"
rmdir /s /q "{staging}"
start "" "{sys.argv[0]}"
del "%~f0"
""")
        return bat
